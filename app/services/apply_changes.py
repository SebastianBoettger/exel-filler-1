from __future__ import annotations
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def save_filled(df: pd.DataFrame, out_dir: str | Path, base_name: str) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d_%H%M")
    out_dir = Path(out_dir)
    out = out_dir / f"{base_name}_filled_{ts}.xlsx"
    df2 = df.drop(columns=[c for c in ["_KEY_"] if c in df.columns])
    df2.to_excel(out, index=False)
    return out

def save_to_path(df: pd.DataFrame, path: str | Path) -> Path:
    path = Path(path)
    df2 = df.drop(columns=[c for c in ["_KEY_"] if c in df.columns])
    df2.to_excel(path, index=False)
    return path

def save_in_place(df: pd.DataFrame, path: str | Path, sheet_name: str, make_backup: bool = True) -> Path:
    path = Path(path)

    if make_backup and path.exists():
        ts = datetime.now().strftime("%Y-%m-%d_%H%M")
        backup = path.with_name(f"{path.stem}_backup_{ts}{path.suffix}")
        backup.write_bytes(path.read_bytes())

    wb = load_workbook(path)

    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    ws = wb.create_sheet(sheet_name, 0)

    df2 = df.drop(columns=[c for c in ["_KEY_"] if c in df.columns])
    for r in dataframe_to_rows(df2, index=False, header=True):
        ws.append(r)

    wb.save(path)
    return path
